#!/usr/bin/env python
# -*- coding: utf-8 -*-

from lib.func_txy import filter_image_paths
from lib.interface import ali1688_multi_image_search, alibaba_multi_image_search, ali1688_similar_offer_urls

import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font

import PySimpleGUI as sg

import os
import sys
import threading
import time
import ctypes

popup_event = threading.Event()

def alert_popup(message, title):
    hwnd = ctypes.windll.user32.GetForegroundWindow()
    ctypes.windll.user32.SetForegroundWindow(hwnd)
    ctypes.windll.user32.FlashWindow(hwnd, 12)

    sg.popup(message, title=title, keep_on_top=True)

def generate_excel(image_paths, interface_choice, window):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "图片+识图搜索结果链接+相似商品链接"
    ws["A1"] = "图片"
    ws["B1"] = "识图搜索结果链接（24小时后无效）"
    ws.column_dimensions["A"].width = 100
    ws.column_dimensions["B"].width = 35
    link_font = Font(color="0563C1", underline="single")
    row = 2

    def progress_callback(success, fail):
        window.write_event_value("progress", (success, fail))

    def popup_callback(str="平台管得比较严，如果遇到登录或者滑块，先帮忙处理"):
        window.write_event_value("popup", str)
        popup_event.wait()
        popup_event.clear()

    if interface_choice == "1688":
        image_search_interface = ali1688_multi_image_search
    elif interface_choice == "Alibaba":
        image_search_interface = alibaba_multi_image_search
    else:
        print(f"\nunknown interface choice {interface_choice}\n" )
        window.write_event_value("exception", "未知的识图搜索接口")
        return

    result_dict = image_search_interface(image_paths, progress_callback, popup_callback)

    for image_path, pair in result_dict.items():
        image = OpenpyxlImage(image_path)
        image.dpi = (300, 300)
        image.width = 100
        image.height = 100
        ws.row_dimensions[row].height = 100
        ws.add_image(image, f"A{row}")

        image_search_url, similar_offer_urls = pair if pair else (None, None)

        if not image_search_url:
            ws[f"B{row}"].value = "识图搜索失败"
        else:
            ws[f"B{row}"].value = image_search_url
            ws[f"B{row}"].hyperlink = image_search_url
            ws[f"B{row}"].style = "Hyperlink"
            ws[f"B{row}"].font = link_font

            if similar_offer_urls:
                col = 3
                for similar_offer_url in similar_offer_urls:
                    header = ws.cell(1, col)
                    header.value = f"相似商品链接{col - 2}"
                    ws.column_dimensions[chr(ord("A") + col - 1)].width = 20

                    data = ws.cell(row, col)
                    data.value = similar_offer_url
                    data.hyperlink = similar_offer_url
                    data.style = "Hyperlink"
                    data.font = link_font
                    col += 1

        row += 1

    if getattr(sys, "frozen", False):
        base_path = os.path.dirname(sys.executable)
    else:
        base_path = os.path.dirname(__file__)

    output_filename = str(os.path.join(base_path, f"{interface_choice}图片+识图搜索结果链接+相似商品链接.xlsx"))

    while (True):
        try:
            wb.save(output_filename)
            window.write_event_value("done", output_filename)
            return
        except PermissionError as e:
            print(f"\n{output_filename} has been opened {e}\n")
            popup_callback(f"{output_filename}文件正处于打开状态，请先关闭文件，否则无法写入")

if __name__ == "__main__":
    if getattr(sys, "frozen", False):
        try:
            base_path = os.path.dirname(sys.executable)
            log_file_path = os.path.join(base_path, "log.txt")
            sys.stderr = sys.stdout = open(log_file_path, "w", encoding="utf-8")
        except Exception as e:
            print("\nlet program go but no log\n")

    image_dir = sg.popup_get_folder("请选择图片所在文件夹", title="选择文件夹")

    if not image_dir:
        sg.popup("未选择文件夹，程序将退出", title="错误")
        sys.exit(0)

    image_paths = filter_image_paths(image_dir)

    if len(image_paths) == 0:
        sg.popup("所选文件夹中没有有效的图片文件，程序将退出", title="错误")
        sys.exit(0)

    interface_choice = sg.popup("请选择识图搜索接口", title="选择接口", custom_text=("1688", "Alibaba"))

    if interface_choice != "1688" and interface_choice != "Alibaba":
        sg.popup("未选择有效接口，程序将退出", title="错误")
        sys.exit(0)

    success = 0
    fail = 0
    total = len(image_paths)

    layout = [
        [sg.Text(f"处理{0}/{total}, 成功{0}个，失败{0}个", key="text")],
        [sg.ProgressBar(total, orientation="h", size=(50, 20), key="progress")]
    ]
    window = sg.Window("处理进度条", layout, finalize=True)

    generate_excel_thread = threading.Thread(target=generate_excel, args=(image_paths, interface_choice, window), daemon=True)
    generate_excel_thread.start()

    start = time.time()

    while True:
        event, values = window.read()

        if event == sg.WIN_CLOSED:
            window.close()
            break
        elif event == "popup":
            alert_popup(values["popup"], "提示")
            popup_event.set()
        elif event == "progress":
            end = time.time()
            success, fail = values["progress"]
            window["text"].update(f"处理{success + fail}/{total}, 成功{success}个, 失败{fail}个, 平均耗时{(end - start) / (success + fail):.3f}s")
            window["progress"].update_bar(success + fail, total)
        elif event == "done":
            output_filename = values["done"]
            alert_popup(f"处理完成！成功: {success}, 失败: {fail}\n结果已保存到 {output_filename}", "完成")
            break
        elif event == "exception":
            exception_name = values["exception"]
            alert_popup(f"{exception_name}, 程序终止", "异常")
            break

    window.close()